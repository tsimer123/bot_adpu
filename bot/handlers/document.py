import pandas as pd
import re
import xlsxwriter
import openpyxl as op
from io import BytesIO
from aiogram import types
from bot.create_bot import dp, bot
from sql.engine import engine
from sqlalchemy.types import Text

db = engine
async def download_document(message: types.Message) -> None:
    output = BytesIO()
    await message.document.download(destination = output)
    df2 = pd.read_excel(BytesIO(filtration(output)), dtype=object)
    #df2.columns[0] = df2.columns[0].astype(str)
    #df2.columns[0] = (df2.columns[0]).replace(' ', '')
    if df2.columns[0] == 'msisdn':
        df2['msisdn'] = df2['msisdn'].astype(str)
        df2['msisdn'] = df2['msisdn'].str.strip()
        conn = db.connect()
        df2.to_sql('simki', con=conn, if_exists='replace', dtype={"msisdn": Text()}) 
        sql_query = pd.read_sql("SELECT msisdn, COALESCE(iccid, 'Нет данных') as iccid, COALESCE(operator, 'Нет данных') as operator, COALESCE(ip, 'Нет данных') as ip, COALESCE(apn, 'Нет данных') as apn, COALESCE(apnusername, 'Нет данных') as apnusername, COALESCE(password, 'Нет данных') as password FROM simki LEFT JOIN simcards USING (msisdn) order by simki.index;", con=conn)
        df = pd.DataFrame(sql_query, columns = ['msisdn', 'iccid', 'operator', 'ip', 'apn', 'apnusername', 'password'])
        conn.close()
        await bot.send_document(message.chat.id, ('response.xlsx', fit(df)))
    elif df2.columns[0] == 'imsi':
        df2['imsi'] = df2['imsi'].astype(str)
        df2['imsi'] = df2['imsi'].str.strip()
        df2 = df2.rename(columns={'imsi': 'iccid'})
        conn = db.connect()
        df2.to_sql('simki', con=conn, if_exists='replace', dtype={"imsi": Text()}) 
        sql_query = pd.read_sql("SELECT iccid, COALESCE(msisdn, 'Нет данных') as msisdn, COALESCE(operator, 'Нет данных') as operator, COALESCE(ip, 'Нет данных') as ip, COALESCE(apn, 'Нет данных') as apn, COALESCE(apnusername, 'Нет данных') as apnusername, COALESCE(password, 'Нет данных') as password FROM simki LEFT JOIN simcards USING (iccid) order by simki.index;", con=conn)
        df = pd.DataFrame(sql_query, columns = ['iccid', 'msisdn', 'operator', 'ip', 'apn', 'apnusername', 'password'])
        conn.close()
        await bot.send_document(message.chat.id, ('response.xlsx', fit(df)))
    elif df2.columns[0] == 'iccid':
        df2['iccid'] = df2['iccid'].astype(str)
        df2['iccid'] = df2['iccid'].str.strip()
        conn = db.connect()
        df2.to_sql('simki', con=conn, if_exists='replace', dtype={"iccid": Text()}) 
        sql_query = pd.read_sql("SELECT iccid, COALESCE(number_tel, 'Нет данных') as number_tel, COALESCE(apn, 'Нет данных') as apn, COALESCE(ip, 'Нет данных') as ip, COALESCE(state, 'Нет данных') as state, activity, COALESCE(traffic, 'Нет данных') as traffic, COALESCE(operator, 'Нет данных') as operator, COALESCE(imei, 'Нет данных') as imei FROM simki LEFT JOIN sims USING (iccid) where sims.state_in_lk='present' order by simki.index;", con=conn)
        df = pd.DataFrame(sql_query, columns = ['iccid', 'number_tel', 'apn', 'ip', 'state', 'activity', 'traffic', 'operator', 'imei'])
        conn.close()
        await bot.send_document(message.chat.id, ('response.xlsx', fit(df)))
    elif df2.columns[0] == 'tel':
        df2['tel'] = df2['tel'].astype(str)
        df2['tel'] = df2['tel'].str.strip()
        df2 = df2.rename(columns={'tel': 'number_tel'})
        conn = db.connect()
        df2.to_sql('simki', con=conn, if_exists='replace', dtype={"number_tel": Text()}) 
        sql_query = pd.read_sql("SELECT number_tel as tel, COALESCE(iccid, 'Нет данных') as iccid, COALESCE(apn, 'Нет данных') as apn, COALESCE(ip, 'Нет данных') as ip, COALESCE(state, 'Нет данных') as state, activity, COALESCE(traffic, 'Нет данных') as traffic, COALESCE(operator, 'Нет данных') as operator, COALESCE(imei, 'Нет данных') as imei FROM simki LEFT JOIN sims USING (number_tel) where sims.state_in_lk='present' order by simki.index;", con=conn)
        df = pd.DataFrame(sql_query, columns = ['tel', 'iccid', 'apn', 'ip', 'state', 'activity', 'traffic', 'operator', 'imei'])
        conn.close()
        await bot.send_document(message.chat.id, ('response.xlsx', fit(df)))
    elif df2.columns[0] == 'ip':
        df2['ip'] = df2['ip'].astype(str)
        df2['ip'] = df2['ip'].str.strip()
        conn = db.connect()
        df2.to_sql('simki', con=conn, if_exists='replace', dtype={"ip": Text()}) 
        sql_query = pd.read_sql("SELECT ip, COALESCE(msisdn, 'Нет данных') as msisdn, COALESCE(iccid, 'Нет данных') as iccid, COALESCE(operator, 'Нет данных') as operator, COALESCE(apn, 'Нет данных') as apn, COALESCE(apnusername, 'Нет данных') as apnusername, COALESCE(password, 'Нет данных') as password FROM simki LEFT JOIN simcards USING (ip) order by simki.index;", con=conn)
        df = pd.DataFrame(sql_query, columns = ['ip', 'msisdn', 'iccid', 'operator', 'apn', 'apnusername', 'password'])
        conn.close()
        await bot.send_document(message.chat.id, ('response.xlsx', fit(df)))
    else:
        await message.reply(
                    f"<b>Обрабатываются только файлы с названиями первых столбцов ip, iccid, msisdn, tel или imsi</b>",
                    parse_mode = 'HTML'
        )
def fit (df):
    output3 = BytesIO()
    writer = pd.ExcelWriter(output3, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Лист1')
    workbook = writer.book
    worksheet = writer.sheets['Лист1']
    worksheet.autofit()
    (max_row, max_col) = df.shape
    column_settings = [{'header': column} for column in df.columns]
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings, 'style': 'Table Style Medium 11' })
    workbook.close()
    document = output3.getvalue()
    return document
def filtration (output):
    output2 = BytesIO()
    output.getvalue()
    excel_doc = op.open(output, data_only=True)
    sheetnames = excel_doc.sheetnames
    sheet = excel_doc[sheetnames[0]]
    rowList = []
    i=1
    while sheet.cell(row = i, column = 1).value is not None:
        a = sheet.cell(row = i, column = 1).value
        if (type(a)) == float:
            rowList.append(i)
        i += 1
    for i in reversed(rowList):
        sheet.delete_rows(i)
    excel_doc.save(output2)
    document = output2.getvalue()
    return document
